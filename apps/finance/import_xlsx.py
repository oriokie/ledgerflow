"""Excel (.xlsx) bulk import for Bills and Recurring transactions.

Not the transaction/statement importer — that stays CSV-only (see
`import_csv.py`), because a bank export is a machine-generated dump of what
already happened. This is for a human filling in a spreadsheet by hand to
enter many bills or recurring charges at once, which is why payee/category/
account are resolved by *name* (case-insensitive) rather than by id: nobody
hand-typing a spreadsheet knows a UUID, and amounts are typed in major units
("42.50"), not the signed minor-unit convention a statement export uses.

Two independent parsers, not one generic "row importer": Bills and Recurring
have different required columns and different FK-resolution targets, and
forcing them through one generic shape would just relocate the per-model
logic into a pile of `if model == "bill"` branches. Both follow `import_csv.py`'s
shape closely enough that a reader of one immediately understands the other —
alias-based column resolution, per-row savepoints, all-or-report.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl
from django.db import transaction

from .models import Category, FinancialAccount


class ImportError_(Exception):
    """Named with a trailing underscore to avoid shadowing the builtin."""


@dataclass
class ImportResult:
    created: int = 0
    errors: list[dict] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {"created": self.created, "errors": self.errors}


def _resolve_columns(header: list[str], aliases: dict[str, set[str]]) -> dict[str, int]:
    lower = {str(h).strip().lower(): i for i, h in enumerate(header) if h}
    resolved: dict[str, int] = {}
    for logical, alias_set in aliases.items():
        for alias in alias_set:
            if alias in lower:
                resolved[logical] = lower[alias]
                break
    return resolved


def _cell(row: tuple, cols: dict[str, int], key: str):
    idx = cols.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_date_cell(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())  # a plain "YYYY-MM-DD" string cell


def _parse_amount_minor(value) -> int:
    """Sheets carry major units (e.g. 42.50), unlike the CSV importer's signed
    minor-unit convention — a spreadsheet author writes '42.50', never '4250'."""
    if isinstance(value, (int, float)):
        return round(float(value) * 100)
    return round(float(str(value).replace(",", "").replace("$", "").strip()) * 100)


def _load_sheet(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in (next(rows, None) or [])]
    return header, rows


def _row_is_blank(row: tuple) -> bool:
    return row is None or not any(c not in (None, "") for c in row)


# --------------------------------------------------------------------------- bills
_BILL_ALIASES: dict[str, set[str]] = {
    "name": {"name", "bill", "description"},
    "amount": {"amount", "amount_minor"},
    "currency": {"currency"},
    "due_on": {"due_on", "due_date", "due"},
    "payee": {"payee"},
    "category": {"category"},
    "recurrence_frequency": {"recurrence_frequency", "frequency"},
    "recurrence_interval": {"recurrence_interval", "interval"},
    "notes": {"notes"},
}

BILL_TEMPLATE_HEADER = [
    "name",
    "amount",
    "currency",
    "due_on",
    "payee",
    "category",
    "recurrence_frequency",
    "recurrence_interval",
    "notes",
]
BILL_TEMPLATE_ROWS = [
    ["Rent", 1500.00, "USD", "2026-02-01", "Sunrise Apartments", "Housing", "monthly", 1, ""],
    ["Netflix", 15.49, "USD", "2026-02-05", "Netflix", "Entertainment", "monthly", 1, ""],
]


def bills_template_xlsx() -> bytes:
    """The blank Bills import template, as .xlsx bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(BILL_TEMPLATE_HEADER)
    for row in BILL_TEMPLATE_ROWS:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@transaction.atomic
def import_bills_xlsx(*, file_bytes: bytes) -> ImportResult:
    from . import bills as bills_service
    from .payees import get_or_create_payee

    result = ImportResult()
    header, rows = _load_sheet(file_bytes)
    cols = _resolve_columns(header, _BILL_ALIASES)
    missing = {"name", "amount", "currency", "due_on"} - set(cols)
    if missing:
        raise ImportError_(f"Sheet missing required column(s): {', '.join(sorted(missing))}.")

    for line_no, row in enumerate(rows, start=2):
        if _row_is_blank(row):
            continue
        try:
            name = str(_cell(row, cols, "name")).strip()
            amount_minor = _parse_amount_minor(_cell(row, cols, "amount"))
            currency = str(_cell(row, cols, "currency")).strip().upper()
            due_on = _parse_date_cell(_cell(row, cols, "due_on"))

            payee = None
            payee_name = _cell(row, cols, "payee")
            if payee_name:
                payee, _created = get_or_create_payee(name=str(payee_name).strip())

            # Look-up only, never auto-created: an unresolved name just means
            # this bill has no category, same as leaving it blank on the
            # manual form — not an error, since amount/due_on are still fully
            # usable without it.
            category = None
            category_name = _cell(row, cols, "category")
            if category_name:
                category = Category.objects.filter(name__iexact=str(category_name).strip()).first()

            recurrence_frequency = _cell(row, cols, "recurrence_frequency") or ""
            recurrence_interval_raw = _cell(row, cols, "recurrence_interval")
            recurrence_interval = int(recurrence_interval_raw) if recurrence_interval_raw else 1
            notes = _cell(row, cols, "notes") or ""

            with transaction.atomic():  # per-row savepoint — one bad row must not sink the sheet
                bills_service.create_bill(
                    name=name,
                    amount_minor=amount_minor,
                    currency=currency,
                    due_on=due_on,
                    payee=payee,
                    category=category,
                    recurrence_frequency=str(recurrence_frequency),
                    recurrence_interval=recurrence_interval,
                    notes=str(notes),
                )
            result.created += 1
        except Exception as exc:  # noqa: BLE001 - report and continue, mirrors import_csv.py
            result.errors.append({"row": line_no, "message": str(exc)})

    return result


# --------------------------------------------------------------------------- recurring
_RECURRING_ALIASES: dict[str, set[str]] = {
    "txn_type": {"txn_type", "type"},
    "account": {"account", "financial_account"},
    "counter_account": {"counter_account", "to_account"},
    "category": {"category"},
    "payee": {"payee"},
    "amount": {"amount", "amount_minor"},
    "currency": {"currency"},
    "frequency": {"frequency"},
    "interval": {"interval"},
    "starts_on": {"starts_on", "start_date", "starts"},
    "ends_on": {"ends_on", "end_date", "ends"},
    "max_occurrences": {"max_occurrences", "occurrences"},
    "memo": {"memo", "notes", "description"},
}

RECURRING_TEMPLATE_HEADER = [
    "txn_type",
    "account",
    "counter_account",
    "category",
    "payee",
    "amount",
    "currency",
    "frequency",
    "interval",
    "starts_on",
    "ends_on",
    "max_occurrences",
    "memo",
]
RECURRING_TEMPLATE_ROWS = [
    ["expense", "Checking", "", "Housing", "Sunrise Apartments", 1500.00, "USD", "monthly", 1,
     "2026-02-01", "", "", "Rent"],
    ["transfer", "Checking", "Savings", "", "", 200.00, "USD", "monthly", 1, "2026-02-01", "", "", "Savings sweep"],
]


def recurring_template_xlsx() -> bytes:
    """The blank Recurring import template, as .xlsx bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(RECURRING_TEMPLATE_HEADER)
    for row in RECURRING_TEMPLATE_ROWS:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@transaction.atomic
def import_recurring_xlsx(*, file_bytes: bytes) -> ImportResult:
    from . import recurring as recurring_service
    from .payees import get_or_create_payee

    result = ImportResult()
    header, rows = _load_sheet(file_bytes)
    cols = _resolve_columns(header, _RECURRING_ALIASES)
    missing = {"txn_type", "account", "amount", "currency", "frequency", "starts_on"} - set(cols)
    if missing:
        raise ImportError_(f"Sheet missing required column(s): {', '.join(sorted(missing))}.")

    for line_no, row in enumerate(rows, start=2):
        if _row_is_blank(row):
            continue
        try:
            txn_type = str(_cell(row, cols, "txn_type")).strip().lower()
            account_name = str(_cell(row, cols, "account")).strip()
            account = FinancialAccount.objects.filter(name__iexact=account_name).first()
            if account is None:
                raise ImportError_(f"No account named {account_name!r}.")

            counter_account = None
            counter_name = _cell(row, cols, "counter_account")
            if counter_name:
                counter_account = FinancialAccount.objects.filter(name__iexact=str(counter_name).strip()).first()
                if counter_account is None:
                    raise ImportError_(f"No account named {str(counter_name).strip()!r}.")

            # Look-up only, never auto-created. Unlike Bills, an unresolved
            # name on an income/expense row isn't silently dropped — it's
            # left None here and create_recurring_transaction itself raises
            # RecurringError for that combination, exactly as it would for
            # any other caller that omitted a required category.
            category = None
            category_name = _cell(row, cols, "category")
            if category_name:
                category = Category.objects.filter(name__iexact=str(category_name).strip()).first()

            payee = None
            payee_name = _cell(row, cols, "payee")
            if payee_name:
                payee, _created = get_or_create_payee(name=str(payee_name).strip())

            amount_minor = _parse_amount_minor(_cell(row, cols, "amount"))
            currency = str(_cell(row, cols, "currency")).strip().upper()
            frequency = str(_cell(row, cols, "frequency")).strip().lower()
            interval_raw = _cell(row, cols, "interval")
            interval = int(interval_raw) if interval_raw else 1
            starts_on = _parse_date_cell(_cell(row, cols, "starts_on"))
            ends_on_raw = _cell(row, cols, "ends_on")
            ends_on = _parse_date_cell(ends_on_raw) if ends_on_raw else None
            max_occurrences_raw = _cell(row, cols, "max_occurrences")
            max_occurrences = int(max_occurrences_raw) if max_occurrences_raw else None
            memo = _cell(row, cols, "memo") or ""

            with transaction.atomic():  # per-row savepoint — one bad row must not sink the sheet
                recurring_service.create_recurring_transaction(
                    txn_type=txn_type,
                    financial_account=account,
                    counter_account=counter_account,
                    category=category,
                    payee=payee,
                    amount_minor=amount_minor,
                    currency=currency,
                    frequency=frequency,
                    interval=interval,
                    starts_on=starts_on,
                    ends_on=ends_on,
                    max_occurrences=max_occurrences,
                    memo=str(memo),
                )
            result.created += 1
        except Exception as exc:  # noqa: BLE001 - report and continue, mirrors import_csv.py
            result.errors.append({"row": line_no, "message": str(exc)})

    return result
